"""解析器基类:负责把各种渠道账单 / PMS 导出的表格统一成 ParsedOrder。

设计要点:
- 支持 CSV(自动识别 utf-8-sig / gbk / utf-8)与 Excel(.xlsx)
- 表头按"别名词典"模糊映射,渠道改版换列名也能容错
- 日期、金额、状态做容错解析(渠道账单格式五花八门)
"""
from __future__ import annotations

import csv
import re
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from ..models import ParsedOrder

# ---------------- 基础工具 ----------------

def parse_amount(v) -> float:
    """解析金额:兼容 '¥1,234.56'、'1,234.56元'、'—'、'(123)' 负数、None、数字。"""
    if v is None:
        return 0.0
    if isinstance(v, bool):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("¥", "").replace("￥", "").replace("元", "").replace(",", "").replace(" ", "").strip()
    if not s or s in ("-", "—", "--", "/", "null", "none"):
        return 0.0
    m = re.match(r"^\(([\d.]+)\)$", s)  # 会计负数 (123)
    if m:
        return -float(m.group(1))
    try:
        return float(s)
    except ValueError:
        m = re.search(r"-?\d+(\.\d+)?", s)  # 取第一个数字段,如 '15%' -> 15
        return float(m.group(0)) if m else 0.0


_DATE_PATTERNS = [
    (re.compile(r"^\d{4}[-/.]\d{1,2}[-/.]\d{1,2}$"), "%Y-%m-%d"),
    (re.compile(r"^\d{4}年\d{1,2}月\d{1,2}日?$"), "%Y年%m月%d日"),
    (re.compile(r"^\d{8}$"), "%Y%m%d"),
]


def parse_date(v) -> str:
    """解析日期,统一输出 YYYY-MM-DD;失败返回空串。"""
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return ""
    s2 = re.split(r"[ T]", s)[0].strip()  # 截掉 '2024-05-01 12:00:00' 的时间部分
    s3 = re.sub(r"[/.]", "-", s2)  # 统一分隔符: '2025/6/1' -> '2025-6-1'
    for pat, fmt in _DATE_PATTERNS:
        if pat.match(s3):
            try:
                return datetime.strptime(s3, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return ""


def parse_nights(v) -> int:
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    m = re.search(r"\d+", str(v))
    return int(m.group(0)) if m else 0


def _norm_header(h) -> str:
    """表头归一化:去空白 / 下划线 / 括号 / 特殊字符,转小写,用于别名匹配。"""
    if h is None:
        return ""
    s = str(h).strip().lower()
    return re.sub(r"[_\s（）()\[\]【】·:：/\\\-]", "", s)


# ---------------- 解析器基类 ----------------

class BaseParser:
    """子类只需定义 channel / COLUMN_ALIASES / STATUS_MAP,必要时覆写 row_to_order。"""

    channel: str = "base"
    # 是否保留没有订单号的行(PMS 侧直客单没有渠道订单号,必须保留)
    keep_rows_without_order_no: bool = False
    COLUMN_ALIASES: dict = {}   # 规范字段 -> [原始表头别名]
    STATUS_MAP: dict = {
        "confirmed": ["成交", "完成", "已入住", "已离店", "已结算", "正常", "有效",
                      "结算", "finished", "completed", "confirmed", "checked"],
        "cancelled": ["取消", "退订", "退款", "cancelled", "canceled", "cancel"],
        "no_show": ["未入住", "noshow", "no-show", "no show", "未到店", "未到"],
    }

    def __init__(self):
        self._header_map: dict = {}   # 归一化表头 -> 规范字段
        for canon, aliases in self.COLUMN_ALIASES.items():
            for a in aliases:
                self._header_map[_norm_header(a)] = canon

    # ---------- 文件读取 ----------

    def read_table(self, path) -> tuple:
        """读取 CSV / Excel,返回 (表头 list, 数据行 list)。"""
        path = Path(path)
        if path.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
            return self._read_excel(path)
        return self._read_csv(path)

    def _read_csv(self, path: Path):
        raw = path.read_bytes()
        text = None
        for enc in ("utf-8-sig", "gbk", "utf-8"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise ValueError(f"无法识别文件编码: {path.name}")
        rows = [r for r in csv.reader(text.splitlines()) if any((c or "").strip() for c in r)]
        if not rows:
            raise ValueError(f"文件为空: {path.name}")
        return rows[0], rows[1:]

    def _read_excel(self, path: Path):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(c is not None and str(c).strip() for c in row):
                rows.append(["" if c is None else c for c in row])
        wb.close()
        if not rows:
            raise ValueError(f"文件为空: {path.name}")
        return rows[0], rows[1:]

    # ---------- 列映射 ----------

    def map_columns(self, headers) -> dict:
        """返回 {规范字段: 列下标}。先精确匹配,再包含匹配(如 '订单号(携程)' 含 '订单号')。"""
        mapping: dict = {}
        used = set()
        for i, h in enumerate(headers):
            canon = self._header_map.get(_norm_header(h))
            if canon and canon not in mapping:
                mapping[canon] = i
                used.add(i)
        for i, h in enumerate(headers):
            if i in used:
                continue
            nh = _norm_header(h)
            for alias, canon in self._header_map.items():
                if canon in mapping or not alias:
                    continue
                if alias in nh or nh in alias:
                    mapping[canon] = i
                    used.add(i)
                    break
        return mapping

    def _cell(self, row, mapping: dict, field: str):
        i = mapping.get(field)
        if i is None or i >= len(row):
            return None
        return row[i]

    def _map_status(self, v) -> str:
        if v is None:
            return "unknown"
        s = str(v).strip().lower()
        if not s:
            return "unknown"
        for canon, kws in self.STATUS_MAP.items():
            for kw in kws:
                if kw.lower() in s:
                    return canon
        return "unknown"

    # ---------- 行解析(子类可覆写) ----------

    def row_to_order(self, row, mapping) -> Optional[ParsedOrder]:
        order = ParsedOrder(channel=self.channel)
        order.order_no = str(self._cell(row, mapping, "order_no") or "").strip()
        order.guest_name = str(self._cell(row, mapping, "guest_name") or "").strip()
        order.room_type = str(self._cell(row, mapping, "room_type") or "").strip()
        order.nights = parse_nights(self._cell(row, mapping, "nights"))
        order.check_in = parse_date(self._cell(row, mapping, "check_in"))
        order.check_out = parse_date(self._cell(row, mapping, "check_out"))
        order.order_amount = parse_amount(self._cell(row, mapping, "order_amount"))
        order.commission_amount = parse_amount(self._cell(row, mapping, "commission_amount"))
        order.status = self._map_status(self._cell(row, mapping, "status"))
        order.book_time = str(self._cell(row, mapping, "book_time") or "").strip()
        settle = self._cell(row, mapping, "settle_amount")
        if settle not in (None, "") and str(settle).strip() not in ("", "-", "—"):
            order.settle_amount = parse_amount(settle)
        else:
            # 结算金额缺省 = 订单金额 - 佣金
            order.settle_amount = round(order.order_amount - order.commission_amount, 2)
        # 保留原始行,便于调试与审计
        order.raw = {field: row[idx] for field, idx in mapping.items() if idx < len(row)}
        return order

    def analyze(self, path) -> dict:
        """诊断表头识别情况:返回 表头 / 已识别字段 / 未识别列,便于排查格式问题。"""
        headers, rows = self.read_table(path)
        mapping = self.map_columns(headers)
        used = set(mapping.values())
        return {
            "headers": headers,
            "recognized": sorted(mapping.keys()),
            "unrecognized_headers": [h for i, h in enumerate(headers) if i not in used],
        }

    def parse_file(self, path) -> list:
        headers, rows = self.read_table(path)
        mapping = self.map_columns(headers)
        if "order_no" not in mapping:
            diag = self.analyze(path)
            raise ValueError(
                "未识别出订单号列,请检查文件表头是否包含「订单号/订单编号」等。\n"
                f"已识别的字段: {diag['recognized']}\n"
                f"未识别的列: {diag['unrecognized_headers']}"
            )
        orders = []
        for row in rows:
            o = self.row_to_order(row, mapping)
            if o and (o.order_no or self.keep_rows_without_order_no):
                orders.append(o)
        return orders

    def parse_with_mapping(self, path, header_map: dict) -> list:
        """按用户配置的表头映射解析: header_map = {标准字段: 源表头原文}。

        与 parse_file 的区别:列映射由用户显式指定(UI 字段映射/预设),
        不再依赖内部别名猜测 —— 所以任何表头都能对账;未映射的可选字段走默认推导。
        """
        headers, rows = self.read_table(path)
        norm_headers = {_norm_header(h): i for i, h in enumerate(headers)}
        mapping = {}
        for canon, src in (header_map or {}).items():
            if not src:
                continue
            idx = norm_headers.get(_norm_header(src))
            if idx is not None:
                mapping[canon] = idx
        if "order_no" not in mapping:
            raise ValueError(
                "字段映射中缺少「订单号」列,请检查映射配置。\n"
                f"文件表头: {headers}\n映射配置: {header_map}")
        orders = []
        for row in rows:
            o = self.row_to_order(row, mapping)
            if o and (o.order_no or self.keep_rows_without_order_no):
                orders.append(o)
        return orders
