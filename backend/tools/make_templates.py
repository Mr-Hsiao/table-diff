"""生成《酒店对账 Excel 模板包》(引流留资用),输出到 table-diff/templates/。

包含:
1. 渠道账单标准化模板.xlsx —— 把渠道账单整理成工具可识别的格式
2. PMS导出模板.xlsx         —— 引导酒店从 PMS 按此格式导出
3. 差异核对表.xlsx          —— 不会用工具的人先用 Excel VLOOKUP 版
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

OUT = Path(__file__).resolve().parents[2] / "templates"
OUT.mkdir(parents=True, exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="1A73E8")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NOTE_FONT = Font(color="6B7A8D", size=10)
WARN_FONT = Font(color="D64541", size=10)


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def add_note_sheet(wb, title, lines, warns=()):
    ws = wb.create_sheet("使用说明")
    ws.cell(1, 1, title).font = Font(bold=True, size=13)
    for i, line in enumerate(lines, start=3):
        ws.cell(i, 1, line).font = NOTE_FONT
    for j, line in enumerate(warns, start=3 + len(lines) + 1):
        ws.cell(j, 1, line).font = WARN_FONT
    ws.column_dimensions["A"].width = 95


def make_channel_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "渠道账单"
    headers = ["订单号", "入住人", "房型", "间夜", "入住日期", "离店日期",
               "订单金额", "佣金比例", "佣金", "结算金额", "订单状态", "下单时间"]
    ws.append(headers)
    ws.append(["1001", "示例客人", "大床房", 1, "2025-06-01", "2025-06-02",
               388, "15%", 38.8, 349.2, "已成交", "2025-05-25 12:30:00"])
    style_header(ws, len(headers))
    add_note_sheet(wb, "渠道账单标准化模板", [
        "1. 把携程/美团商家后台导出的结算单,整理成此格式(列名近似即可,工具会自动识别)",
        "2. 必填列:订单号、入住人、入住日期、离店日期、订单金额、佣金、结算金额、订单状态",
        "3. 佣金比例和佣金填一个即可;结算金额不填时,自动用 订单金额-佣金 计算",
        "4. 订单状态按实际情况写:已成交/已入住/已取消/未入住 等,工具会自动归类",
        "5. 表格支持 Excel(.xlsx)和 CSV 两种格式",
    ], warns=("提示:此模板用于了解格式;自动对账请用「表对比工具」,可私信了解。"))
    wb.save(OUT / "渠道账单标准化模板.xlsx")


def make_pms_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "PMS导出"
    headers = ["内部单号", "渠道订单号", "客人", "房型", "入住日期", "离店日期",
               "间夜", "房费", "佣金", "实收金额", "来源渠道", "状态"]
    ws.append(headers)
    ws.append(["A001", "1001", "示例客人", "大床房", "2025-06-01", "2025-06-02",
               1, 388, 38.8, 349.2, "携程", "已离店"])
    ws.append(["C001", "", "散客", "大床房", "2025-06-09", "2025-06-10",
               1, 300, 0, 300, "直客", "已离店"])
    style_header(ws, len(headers))
    add_note_sheet(wb, "PMS 导出模板", [
        "1. 从 PMS 导出营业数据,整理成此格式(不同 PMS 的报表列名不同,保留近似列名即可)",
        "2. 关键列:渠道订单号(OTA 订单号)——对账就靠它匹配,直客单可留空",
        "3. 实收金额 = 房费 - 佣金(OTA 单);直客单实收 = 房费",
        "4. 找不到渠道订单号时,工具会按 客人+入住日期+金额 模糊匹配",
    ])
    wb.save(OUT / "PMS导出模板.xlsx")


def make_diff_template():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "渠道账单"
    ws1.append(["订单号", "结算金额"])
    ws1.append([1001, 349.2])
    ws1.append([1002, 698.4])
    ws1.append([1003, 529.2])
    style_header(ws1, 2)

    ws2 = wb.create_sheet("PMS数据")
    ws2.append(["订单号", "实收金额"])
    ws2.append([1001, 349.2])
    ws2.append([1002, 698.4])
    ws2.append([1003, 529.2])
    style_header(ws2, 2)

    ws3 = wb.create_sheet("差异核对")
    ws3.append(["渠道订单号", "渠道结算金额", "PMS实收金额", "差异(渠道-PMS)"])
    ws3.append([1001, "=VLOOKUP(A2,渠道账单!$A:$B,2,FALSE)", "=VLOOKUP(A2,PMS数据!$A:$B,2,FALSE)", "=B2-C2"])
    ws3.append([1002, "=VLOOKUP(A3,渠道账单!$A:$B,2,FALSE)", "=VLOOKUP(A3,PMS数据!$A:$B,2,FALSE)", "=B3-C3"])
    ws3.append([1003, "=VLOOKUP(A4,渠道账单!$A:$B,2,FALSE)", "=VLOOKUP(A4,PMS数据!$A:$B,2,FALSE)", "=B4-C4"])
    style_header(ws3, 4)
    add_note_sheet(wb, "差异核对表(Excel 版)", [
        "1. 把渠道账单粘贴到「渠道账单」表,把 PMS 数据粘贴到「PMS数据」表",
        "2. 在「差异核对」表输入要核对的订单号,公式会自动带出两边金额和差异",
        "3. 差异列不为 0 的就是要对账的订单",
        "4. 说明:这只是基础核对,自动对账(含 No-show 佣金、漏单、单边单检查)请用「表对比工具」",
    ], warns=("提示:订单号较多(几百行)时,公式版会卡;此时建议用自动对账工具。"))
    wb.save(OUT / "差异核对表.xlsx")


if __name__ == "__main__":
    make_channel_template()
    make_pms_template()
    make_diff_template()
    print("模板包已生成到:", OUT)
    for f in sorted(OUT.iterdir()):
        print(" -", f.name, f.stat().st_size, "bytes")
